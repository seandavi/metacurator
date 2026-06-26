"""tables — load supplements (xlsx/docx/pdf/csv) -> SourceTable. SPEC 050. [deterministic]

Mechanical loading only — no interpretation of meaning (that is SPEC 100). The library
extraction (openpyxl/python-docx/pdfplumber/csv) is kept separate from the shared
``raw rows -> SourceTable`` conversion so header detection is one code path across formats.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from .models import SourceProvenance, SourceTable


@dataclass
class Frame:
    """Loaded table representation kept behind ``SourceTable.frame`` (SPEC 050)."""

    columns: list[str]
    records: list[dict[str, Any]] = field(default_factory=list)

    def __len__(self) -> int:
        return len(self.records)


def _detect_header(rows: list[list[Any]]) -> int:
    """Index of the first fully-populated row (SPEC 050).

    The header is the first row whose filled-cell count matches the widest row seen, so a
    sparse title/caption line above the real header is skipped.
    """
    fill_counts = [sum(1 for c in row if c not in (None, "")) for row in rows]
    if not fill_counts:
        return 0
    widest = max(fill_counts)
    return fill_counts.index(widest)


def _clean_headers(header: list[Any]) -> list[str]:
    """Stringify, fill blanks, and disambiguate duplicate header cells."""
    out: list[str] = []
    seen: dict[str, int] = {}
    for j, cell in enumerate(header):
        name = "" if cell is None else str(cell).strip()
        if not name:
            name = f"col{j + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        out.append(name)
    return out


def _rows_to_table(
    raw_rows: list[list[Any]], provenance: SourceProvenance
) -> SourceTable | None:
    """Header-detect + build a SourceTable from raw rows; ``None`` if empty."""
    rows = [r for r in raw_rows if any(c not in (None, "") for c in r)]
    if not rows:
        return None
    header_idx = _detect_header(rows)
    columns = _clean_headers(rows[header_idx])
    records: list[dict[str, Any]] = []
    for data_row in rows[header_idx + 1 :]:
        padded = list(data_row) + [None] * (len(columns) - len(data_row))
        records.append(dict(zip(columns, padded, strict=False)))
    prov = provenance.model_copy(update={"header_row": header_idx})
    return SourceTable(
        frame=Frame(columns=columns, records=records),
        provenance=prov,
        n_rows=len(records),
        n_cols=len(columns),
    )


# -- per-format extraction ---------------------------------------------------


def _load_delimited(path: Path, url: str | None) -> list[SourceTable]:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    with open(path, newline="", encoding="utf-8-sig") as f:
        raw_rows = [list(r) for r in csv.reader(f, delimiter=delimiter)]
    prov = SourceProvenance(file=str(path), url=url)
    table = _rows_to_table(raw_rows, prov)
    return [table] if table else []


def _load_xlsx(path: Path, url: str | None) -> list[SourceTable]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    tables: list[SourceTable] = []
    for sheet in wb.worksheets:
        raw_rows = [list(r) for r in sheet.iter_rows(values_only=True)]
        prov = SourceProvenance(file=str(path), sheet=sheet.title, url=url)
        table = _rows_to_table(raw_rows, prov)
        if table:
            tables.append(table)
    wb.close()
    return tables


def _load_docx(path: Path, url: str | None) -> list[SourceTable]:
    from docx import Document

    doc = Document(str(path))
    tables: list[SourceTable] = []
    for idx, t in enumerate(doc.tables):
        raw_rows = [[cell.text for cell in row.cells] for row in t.rows]
        prov = SourceProvenance(file=str(path), table_index=idx, url=url)
        table = _rows_to_table(raw_rows, prov)
        if table:
            tables.append(table)
    return tables


def _load_pdf(path: Path, url: str | None) -> list[SourceTable]:
    import pdfplumber

    tables: list[SourceTable] = []
    idx = 0
    with pdfplumber.open(str(path)) as pdf:
        for page in pdf.pages:
            for raw_rows in page.extract_tables() or []:
                prov = SourceProvenance(file=str(path), table_index=idx, url=url)
                table = _rows_to_table([list(r) for r in raw_rows], prov)
                if table:
                    tables.append(table)
                idx += 1
    return tables


_LOADERS = {
    ".csv": _load_delimited,
    ".tsv": _load_delimited,
    ".xlsx": _load_xlsx,
    ".docx": _load_docx,
    ".pdf": _load_pdf,
}


def load_tables(path: Path | str, *, url: str | None = None) -> list[SourceTable]:
    """Load every table in a supplement file, with provenance. See SPEC 050."""
    path = Path(path)
    loader = _LOADERS.get(path.suffix.lower())
    if loader is None:
        raise ValueError(f"unsupported supplement type {path.suffix!r} ({path.name})")
    tables = loader(path, url)
    if not tables:
        raise ValueError(f"no tables found in {path.name}")
    return tables
